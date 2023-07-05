from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Load the workbook
workbook = load_workbook('outputTest.xlsx')

# Get the first sheet
sheet = workbook.active

# Iterate through the rows of the sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Get the VIN value from column A
    vin = row[0]
    
    # Create a hyperlink using the VIN value
    hyperlink = '=HYPERLINK("http://example.com/{0}"; "{0}")'.format(vin)
    
    # Get the cell in column A for the current row
    cell = sheet['A{}'.format(row)]
    
    # Set the hyperlink formula
    cell.value = vin
    cell.font = Font(underline='single', color='0563c1')
    cell.hyperlink = hyperlink

# Save the modified workbook
workbook.save('output_modified.xlsx')
